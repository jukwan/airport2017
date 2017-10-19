### web crawling for ifly.com flight information
import schedule
import time
from datetime import datetime, timedelta
import json
import bs4 as bs
import numpy as np
import pandas as pd
import requests
from openpyxl import load_workbook
# start from yesterday. since operation is completed. 

def job():
    date = datetime.strftime(datetime.now() -timedelta(1), '%Y%m%d')
   #date =(datetime.now() - timedelta(1)).strftime('%Y%m%d') # same results
    travels = ['arrairp','depairp'] #'arrairp'

    for travel in travels:
        for airport in ['BOS', 'PVD', 'MHT', 'LAX','SNA','BUR','LGB']: # loop for 4 airports near Boston    
        #   #outcome saving path: same file for same day and airport.
            path = "F:\Dropbox\Flight\Data\output_"+date+travel+airport+".xlsx"
            #path = "D:\Data\output_"+date+travel+airport+".xlsx"
                                
            writer = pd.ExcelWriter(path, engine = 'openpyxl') #  open wirter

            for time in ['0000', '0300', '0600', '0900','1200','1500', '1800','2100']: # loop for time options
                url ="http://www.ifly.com/api/v1/airlines/track?queryby=airport&date="+date+"&hr="+time+"&"+travel+"="+airport # generate the URl for each case
                    #http://www.ifly.com/api/v1/airlines/track?queryby=airport&date=20170829&hr=0900&depairp=BOS
                    #above url is located inspect / network. is is source data for web page.
                print(url) #print url 
                data= requests.get(url).json() # fetch the site - above url has simple Json structure

                    # print (data.status_code)
                    # if "blocked" in data.text:
                    #     print ("we've been blocked")
                    # print (r.headers.get("content-type", "unknown"))
                    #print (data)

                status_json = data['status'] # Check the data condtion
                print('API status:'+ status_json)
                if status_json =="fail":
                    continue # move back to start point of loop
                else:
                    pass # go next line

                # pick the data level dictionary      
                data_1= data['data']
                # narrow down the data which has individual flight data
                flights = data_1['flights']
                #print(flights)

                # Check the number of flights in the data

                if 'numFlights' in data_1: # this is for case where expected # of flight =! actual # of landed Flights
                    print('yes')
                    n= data_1['numFlights']
                    
                else :
                    nn = data_1['message'].split()
                    n= int(nn[0])
                #print(nn[0])
                #print(type(nn[0]))
        
                print(n)

                #make empty list
                flight_n =[]
                airline_name = []
                aircraft_type = [] # ex A320 
                airline_code = []
                arrival_country =[]
                arrival_airport =[]
                arrival_date_UTC = []
                arrival_date = []
                arrival_time_scheduled_UTC = []
                arrival_time_actual_UTC = []
                arrival_time_scheduled = []
                arrival_time_actual = []
                depature_airport = []
                departure_country = []
                departure_date_UTC = []
                departure_date = []
                departure_time_scheduled = []
                departure_time_actual = []
                departure_time_scheduled_UTC = []
                departure_time_actual_UTC = []
                flightstatus = []
                
                # fill the list for each information
                for i in range(0, n):
                    flight_n.append(flights[i]['flightNumber'])
                    airline_code.append(flights[i]['airline']['code'])
                    arrival_airport.append(flights[i]['arrival'][0]['airport']['code'])
                
                    flightstatus.append(flights[i]['flightStatus']['msg'])
                    arrival_date_UTC.append(flights[i]['arrival'][0]['dateTimes'][0]['dateUTC'])
                    arrival_date.append(flights[i]['arrival'][0]['dateTimes'][0]['date'])

                    arrival_time_scheduled_UTC.append(flights[i]['arrival'][0]['dateTimes'][0]['timeUTC'])
                    arrival_time_scheduled.append(flights[i]['arrival'][0]['dateTimes'][0]['time'])

                    departure_date_UTC.append(flights[i]['departure'][0]['dateTimes'][0]['dateUTC'])
                    departure_date.append(flights[i]['departure'][0]['dateTimes'][0]['date'])

                    departure_time_scheduled_UTC.append(flights[i]['departure'][0]['dateTimes'][0]['timeUTC'])
                    departure_time_scheduled.append(flights[i]['departure'][0]['dateTimes'][0]['time'])

                    depature_airport.append(flights[i]['departure'][0]['airport']['code'])

                for i in range(0, n):
                    try:
                        aircraft_type.append(flights[i]['aircraft']['type'])
                    except KeyError :
                        aircraft_type.append('NA')    
                        continue

                for i in range(0, n):
                    try:
                        departure_time_actual.append(flights[i]['departure'][0]['dateTimes'][1]['timeUTC'])
                    except IndexError: # sometimes there is no data for actual departure time
                        departure_time_actual.append('NA')
                        continue

                for i in range(0, n):
                    try:
                        departure_time_actual_UTC.append(flights[i]['departure'][0]['dateTimes'][1]['time'])
                    except IndexError: # sometimes there is no data for actual departure time
                        departure_time_actual_UTC.append('NA')
                        continue

                for i in range(0, n):
                    try:
                        arrival_time_actual.append(flights[i]['arrival'][0]['dateTimes'][1]['timeUTC'])
                    except IndexError: # sometimes there is no data for actual departure time
                        arrival_time_actual.append('NA')
                        continue

                for i in range(0, n):
                    try:
                        arrival_time_actual_UTC.append(flights[i]['arrival'][0]['dateTimes'][1]['time'])
                    except IndexError: # sometimes there is no data for actual departure time
                        arrival_time_actual_UTC.append('NA')
                        continue

                for i in range(0, n):  
                    try:
                        departure_country.append(flights[i]['departure'][0]['airport']['countryId'])
                    except KeyError:
                        departure_country.append('na')
                        continue

                for i in range(0, n) :
                    try: 
                        arrival_country.append(flights[i]['arrival'][0]['airport']['countryId'])
                    except KeyError:
                        arrival_country.append('na')
                        continue

                            
                # print(len(flight_n))
                # print(len(airline_code))
                # print(len(arrival_time_actual))
                # print(arrival_time_actual)
                # print(departure_time_actual)
                # print((flightstatus))

                #switch to data frame for future analysis
                df = pd.DataFrame({"Flight_n":flight_n, "Airline":airline_code, \
                    "Arrival_airport":arrival_airport, "Arrival_date":arrival_date, "Arrival_scheduled":arrival_time_scheduled, \
                    "Arrival_acutal":arrival_time_actual, \
                    "Arrival_date_UTC":arrival_date_UTC, "Arrival_scheduled_UTC":arrival_time_scheduled_UTC , \
                    "Arrival_acutal_UTC":arrival_time_actual_UTC, \
                    "Arrival_Country": arrival_country, \
                    "Origin_airport":depature_airport, "Origin_country":departure_country, \
                    "Departure_date":departure_date, "Departure_time_scheduled":departure_time_scheduled, \
                    "Departure_time_actual":departure_time_actual,\
                    "Departure_date_UTC":departure_date_UTC, "Departure_time_scheduled_UTC":departure_time_scheduled_UTC, \
                    "Departure_time_actual_UTC":departure_time_actual_UTC,
                     "Status":flightstatus , "Aircraft": aircraft_type})

                # re arrange the order of variables since above dataframe give alphabet ordered dataframe
                frame = df[["Airline", "Flight_n", "Aircraft", \
                    "Origin_airport", "Origin_country", "Departure_date", "Departure_time_scheduled", "Departure_time_actual", \
                    "Departure_date_UTC", "Departure_time_scheduled_UTC", "Departure_time_actual_UTC", \
                    "Arrival_airport", "Arrival_Country", "Arrival_date", "Arrival_scheduled", "Arrival_acutal", \
                    "Arrival_date_UTC", "Arrival_scheduled_UTC", "Arrival_acutal_UTC", \
                    "Status" ]]     
                

                ## save to the excel file appending!!
                
                # if time == "0000" :
                #     frame.to_excel(writer, index = False )
                #     writer.save() 
                
                # else :
                try :
                    wb = load_workbook(path)
                    sheet = wb.worksheets[0]
                    curr_count = sheet.max_row 
                    frame.to_excel(writer, startrow = curr_count, header= None, index = False)
                    writer.save()

                except FileNotFoundError :
                    frame.to_excel(writer, index = False )
                    writer.save()
                    pass
                                        
                    
schedule.every().day.at("17:34").do(job)

while True:
    schedule.run_pending()
    time.sleep(1)