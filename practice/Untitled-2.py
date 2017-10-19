import json
import bs4 as bs
import numpy as np
import pandas as pd
import requests
from openpyxl import load_workbook
date = '20170830'
travels = 'arrairp'#,'depairp']
path = "F:\Dropbox\Flight\Data\output_"+date+travels+"aaa.xlsx"

writer = pd.ExcelWriter(path, engine = 'openpyxl') # ? wirte a inital empty file.
# for time in ['0000','1111']:

#     writer = pd.ExcelWriter(path, engine = 'openpyxl') # ? wirte a inital empty file.
#     wb= load_workbook(paht, use_iterator =True)
#     shhet = wb.worksheets[0]
#     curr_count = sheet.max_row
#     if time == "0000" :
#                 frame.to_excel(writer, index = False )
            
#     else :
#         frame.to_excel(writer, startrow = curr_count +1, header= None, index = False)
#     writer.save()